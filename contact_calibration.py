import numpy as np
from scipy.optimize import minimize
from openpyxl import load_workbook, Workbook
from os import path
from math import floor, ceil

contact_base_path = path.join('prem_data')

def letter_to_int(letter):
    alphabet = list('ABCDEFGHIJKLMNOPQRSTUVWXYZ')
    return alphabet.index(letter) + 1

def read_age_pyramid(country):
    # workbook reading
    sheet_path = path.join('country_data', 'age_pyramids', 'formated_data.xlsx')
    wb = load_workbook(sheet_path, read_only=True)
    sheet = wb.get_sheet_by_name('constant')

    if country == "None":
        region = "World"
    else:
        region = country
    age_pyramid = {}
    for line in range(sheet.max_row + 1)[2:]:
        if sheet.cell(row=line, column=1).value == region:
            for col_index in range(4, 21):
                cat_name = sheet.cell(row=1, column=col_index).value.encode("utf-8")
                value = float(sheet.cell(row=line, column=col_index).value)
                age_pyramid[cat_name] = value

    # normalise the vector so it sums to 1.0. It now contains proportions
    s = sum(age_pyramid.values())
    for key, value in age_pyramid.iteritems():
        age_pyramid[key] /= s
    return age_pyramid

def read_matrix(contact_type, country):
    assert contact_type in ['all_locations', 'home', 'other_locations', 'school', 'work']

    # preliminary work to know which file to open
    index = letter_to_int(country[0])
    if index < 13:  # first letter of the country is among ABCDEFGHIJKL
        string_to_add = '_1'
    elif index > 13:  # first letter of the country is among NOPQRSTUVWXYZ
        string_to_add = '_2'
    else:  # first letter is M
        if country == 'Mozambique':
            string_to_add = '_2'
        else:
            string_to_add = '_1'

    # workbook reading
    filename = 'MUestimates_' + contact_type + string_to_add
    sheet_path = path.join(contact_base_path, filename + '.xlsx')
    wb = load_workbook(sheet_path, read_only=True)
    sheet = wb.get_sheet_by_name(country)

    # matrix creation
    row_index = 0
    if string_to_add == '_1': # there is an extra line in the spreadsheet
        row_index = 1

    matrix = np.zeros((16, 16))
    for i in range(16):
        row_index += 1
        col_index = 0
        for j in range(16):
            col_index += 1
            matrix[i, j] = sheet.cell(row=row_index, column=col_index).value

    return matrix

def extract_sub_matrix(matrix, age_min, age_max):
    """
    For some places (work, school), we only need to work with the information about a fixed age category
    :param matrix: numpy array
    :param age_min: float
    :param age_max: float
    :return: numpy array
    """
    low_bound = 5.*ceil(age_min/5.)  # e.g.  18 -> 20    20 -> 20
    index_min = int(low_bound / 5.)

    high_bound = 5.*ceil(age_max/5.)  # e.g.  18 -> 20    20 -> 20
    index_max = int(high_bound / 5.) - 1

    relevant_x_cat = range(index_min+1, index_max+1)
    relevant_x_cat = ["X_" + str(indice) for indice in relevant_x_cat]
    return {'matrix': matrix[index_min:index_max, index_min:index_max], 'relevant_x_cat': relevant_x_cat}

def age_preference_function(age_difference, sd):
    """
    Given the age difference between two individuals, computes the relative probability of contact.
    Reference: no age difference.
    location_type is either "school" or "workplace"
    """
    ratio = np.exp(-age_difference ** 2 / (2 * sd ** 2))
    return ratio

def calibrate_param_for_age_preference(matrix, age_pyramid, relevant_x_cat):
    """
    Given a contact rate matrix as reported in Prem et al and an age-pyramid for the background population,
    we estimate the standard deviation used in the age-preference function.
    :return: the calibrated standard deviation
    """

    def distance_function(sd):
        sq_dist = 0.
        for i in range(matrix.shape[0]):
            if matrix[i, i] > 0:  # if the diagonal item is 0, we do nothing
                for j in range(matrix.shape[0]):
                    if i != j:
                        data_ratio = matrix[i, j] / matrix[i, i]
                        age_dif = 5.*abs(j-i)
                        model_ratio = age_preference_function(age_dif, sd)
                        age_cat_key_j = relevant_x_cat[j]
                        age_cat_key_i = relevant_x_cat[i]
                        population_ratio = age_pyramid[age_cat_key_j] / age_pyramid[age_cat_key_i]  # P_j / P_i
                        sq_dist += (data_ratio - population_ratio * model_ratio) ** 2
        return sq_dist
    sd_0 = [5.]
    bnds = [(0.001, 100.)]
    res = minimize(fun=distance_function, x0=sd_0, method='SLSQP', bounds=bnds)  #only L-BFGS-B, TNC and SLSQP for bounded problems
    best_param = res['x'][0]
    distance = res['fun'][0]

    if best_param == bnds[0][0]:  # rerun optimisation with restricted range
        bnds = [(0.001, 10.)]
        res = minimize(fun=distance_function, x0=sd_0, method='SLSQP',
                       bounds=bnds)  # only L-BFGS-B, TNC and SLSQP for bounded problems
        best_param = res['x'][0]
        distance = res['fun'][0]

    # # linear model
    # def distance_function_linear(slope):
    #     sq_dist = 0.
    #     for i in range(matrix.shape[0]):
    #         if matrix[i, i] > 0:  # if the diagonal item is 0, we do nothing
    #             for j in range(matrix.shape[0]):
    #                 if i != j:
    #                     data_ratio = matrix[i, j] / matrix[i, i]
    #                     age_dif = 5. * abs(j - i)
    #                     model_ratio = 1. + age_dif * slope
    #                     sq_dist += (data_ratio - model_ratio) ** 2
    #     return sq_dist
    #
    # slope_0 = [-1./100.]
    # bnds = [(-1./50., 0.)]
    # res = minimize(fun=distance_function_linear, x0=slope_0, method='SLSQP',
    #                bounds=bnds)  # only L-BFGS-B, TNC and SLSQP for bounded problems
    # slope = res['x'][0]
    # distance_linear = res['fun'][0]
    #
    # if distance_linear < distance:
    #     best_param = slope
    print best_param

    return best_param

def get_country_list():
    # countries workbook reading
    filename = 'countries'
    sheet_path = path.join('country_data', filename + '.xlsx')
    wb = load_workbook(sheet_path, read_only=True)
    sheet = wb.get_sheet_by_name('Sheet1')

    countries = []
    for line in range(sheet.max_row + 1)[1:]:
        if line > 1:
            countries.append(sheet.cell(row=line, column=1).value.encode("utf-8"))
    return countries

def calibrate_param_for_all_countries():
    countries = get_country_list()
    sds = {'school': {}, 'work': {}, 'other_locations':{}}
    for country in countries:
        print country
        age_pyramid = read_age_pyramid(country)
        for contact_type in ['other_locations']:
            if len(age_pyramid) == 0:
                sds[contact_type][country] = None
                print "!!!!!!!! " + country
            else:
                matrix = read_matrix(contact_type, country)
                age_min, age_max = 0., 80.
                if contact_type == 'school':
                    age_min, age_max = 5., 18.
                elif contact_type == 'work':
                    age_min, age_max = 18., 65.
                elif contact_type == 'other_locations':
                    age_min, age_max = 0., 120.
                submatrix = extract_sub_matrix(matrix=matrix, age_min=age_min, age_max=age_max)
                matrix, relevant_x_cat = submatrix['matrix'], submatrix['relevant_x_cat']
                sd = calibrate_param_for_age_preference(matrix, age_pyramid, relevant_x_cat)
                sds[contact_type][country] = sd
    return sds

def write_estimate_to_countries_spreadsheet(estimate_dict, param_name):
    wb = Workbook()
    ws = wb.active
    for key, value in estimate_dict.iteritems():
        ws.append([key, value])

    sheet_path = path.join(contact_base_path, param_name + '.xlsx')
    wb.save(sheet_path)

if __name__ == "__main__":
    estimate_dict = calibrate_param_for_all_countries()
    #write_estimate_to_countries_spreadsheet(estimate_dict['school'], 'param_agepref_school')
    #write_estimate_to_countries_spreadsheet(estimate_dict['work'], 'param_agepref_work')
    write_estimate_to_countries_spreadsheet(estimate_dict['other_locations'], 'sd_agepref_other_locations')

