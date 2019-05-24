from os import path, makedirs
from openpyxl import load_workbook
from numpy import genfromtxt, mean, linspace, asarray, savetxt, array, zeros
from math import floor, ceil, exp
from curve import scale_up_function
from copy import deepcopy
import spreadsheet
import contact_calibration
import copy
from toolkit import lhs_sampler
from scipy.stats.distributions import beta, uniform, triang

def read_sheet(file):
    """
    param file: the path to the file
    return: the sheet object
    """
    wb = load_workbook(file, read_only=True)
    sheet = wb['constant']
    return sheet

def sheet_to_dict(sheet):
    """
    Create a dictionary with the parameters contained in sheet
    param sheet: a sheet object obtained from read_sheet
    return: a dictionary {'par_name'=value, ...}
    """
    params = {}
    for line in range(sheet.max_row+1)[1:]:
        par_name = sheet.cell(row=line, column=1).value
        if par_name is None:
            break
        par_name = par_name.encode('ascii', 'ignore')
        par_val = sheet.cell(row=line, column=2).value
        par_type = sheet.cell(row=line, column=3).value
        if par_type == 'integer':
            par_val = int(par_val)
        elif par_type == 'boolean':
            par_val = (par_val == 1.0)
        elif par_type == 'float':
            par_val = float(par_val)
        elif par_type == 'string' or par_type == 'character':
            par_val = str(par_val)
        params[par_name] = par_val
    return params

def get_iso3(country):
    """
    return the iso3 code of a country
    """
    if country == 'India':
        return 'IND'
    elif country == 'China':
        return 'CHN'

    base_path = path.join('country_data')
    sheet_path = path.join(base_path, 'iso3.xlsx')
    if path.isfile(sheet_path):
        sheet = read_sheet(sheet_path)
        par_dict = sheet_to_dict(sheet)
        iso3 = None
        for key, val in par_dict.iteritems():
            if country.lower() in val.encode('utf8').lower():
                iso3 = key
    else:
        print "Spreadsheet iso3.xlsx does not exist"
    return iso3

def get_agecategory(age, data='pyramid'):
    # data is either 'pyramid' or 'prem'
    max_age = {'pyramid': 80., 'prem': 75.}
    last_cat = {'pyramid': 'X_17', 'prem': 'X_16'}
    if age > max_age[data]:
        return last_cat[data]
    else:
        cat_ind = int(floor(1. + age / 5.))
        return 'X_' + str(cat_ind)

class data:
    """
    Object that stores all of the data found in spreadsheets
    """
    def __init__(self, country=None, calibration_params=None, uncertainty_params=None):
        self.country = country  # only for the purpose of multi-country analysis
        self.calibration_params = calibration_params
        self.uncertainty_params = uncertainty_params
        self.sampled_params = {}
        self.console = {}
        self.common_parameters = {}
        self.data_from_sheets = {}
        self.scale_up_functions = {}
        self.scenario_names = []
        self.scenarios = {}
        self.age_pyramid = {}
        self.siler_params = {}
        self.contact_rates_matrices = {}
        self.prem_contact_rate_functions = {}  # from Prem
        self.birth_numbers_function = None

        self.pool_of_life_durations = []
        self.activation_times_dic = None
        self.sd_agepref_work = {}

        self.read_all_data()
        self.calculate_scale_up_functions()

        self.process_checkpoints()
        self.workout_timeseries()
        self.workout_universal_methods()
        self.workout_birth_rates()

    def create_output_directory(self):
        dir_path = path.join('outputs', self.console['project_name'])
        if not path.exists(dir_path):
            print "Creating directory " + dir_path
            makedirs(dir_path)
        else:
            print "Directory " + dir_path + " already exists. No creation needed."

    def read_all_data(self):
        """
        Reads the data from the different spreadsheets and populates the attributes of the data object
        """
        base_path = path.join('spreadsheets')

        # read the console and the common parameters
        for sheet_name in ['console', 'common_parameters']:
            sheet_path = path.join(base_path, sheet_name + '.xlsx')
            if path.isfile(sheet_path):
                sheet = read_sheet(sheet_path)
                par_dict = sheet_to_dict(sheet)
                setattr(self, sheet_name, par_dict)

            else:
                print "Spreadsheet " + sheet_name + " does not exist"

        # adjust the country value in case of multi-country analysis
        if self.country is not None:
            self.console['country'] = self.country
            self.console['project_name'] = self.console['project_name'] + "_" + self.country

        self.create_output_directory()

        # Overwrite parameters with country-specific parameter values
        if self.console['country'] is not None:
            sheet_path = path.join(base_path, 'country_parameters.xlsx')
            if path.isfile(sheet_path):
                sheet = read_sheet(sheet_path)
                self.read_country_parameters(sheet)
            else:
                print "Spreadsheet " + sheet_name + " does not exist"

        # read the scenario-specific parameters
        if self.console['running_mode'] not in ['run_ks_based_calibration', 'run_lhs_calibration']:  # normal manual run. We read scenario-specific spreadsheets
            for par_name, par_val in self.console.iteritems():
                if 'scenario_' in par_name and par_val:
                    sheet_path = path.join(base_path, par_name + '.xlsx')
                    if path.isfile(sheet_path):
                        sheet = read_sheet(sheet_path)
                        par_dict = sheet_to_dict(sheet)
                        self.scenarios[par_name] = par_dict
                        self.scenario_names.append(par_name)
                    else:
                        print "Spreadsheet " + par_name + " does not exist"
        elif self.console['running_mode'] == 'run_ks_based_calibration':  # We automatically generate scenarios
            for param_name, param_vals in self.calibration_params.iteritems():
                for param_val in param_vals:
                    scenario_name = "calib_" + str(round(param_val, 5))
                    scenario_name = scenario_name.replace('.', '_')
                    self.scenarios[scenario_name] = {param_name: param_val,
                                                     'scenario_title': param_name + "_" + str(round(param_val, 5))}
                    self.scenario_names.append(scenario_name)
        else:  # This is a LHS calibration. We automatically generate scenarios
            self.draw_lhs_parameters()
            self.write_lhs_parameters()
            for i_sample in range(self.console['n_lhs_paramsets']):
                scenario_name = "lhs_sample_" + str(i_sample)
                self.scenarios[scenario_name] = {'scenario_title': 'lhs_' + str(i_sample)}
                for param in self.uncertainty_params.keys():
                    self.scenarios[scenario_name][param] = self.sampled_params[param][i_sample]
                self.scenario_names.append(scenario_name)
        # # read the Mossong contact-rates data
        # sheet_path = path.join(base_path, 'mossong_contact_rates.xlsx')
        # if path.isfile(sheet_path):
        #     sheet = read_sheet(sheet_path)
        #     self.process_mossong_contact_rates(sheet)
        # else:
        #     print "Spreadsheet mossong_contact_rates does not exist"
        # self.create_contact_rate_function()

        # read the pool of life durations from a csv file
        if self.console['country'] == "None":
            sheet_path = path.join(base_path, 'pool_of_life_durations.csv')
        else:
            iso3 = get_iso3(self.console['country'])
            sheet_path = path.join('country_data', 'ages_at_death', iso3 + '.csv')
        if path.isfile(sheet_path):
            ages_at_death = genfromtxt(sheet_path, delimiter=',')
        else:
            print "Spreadsheet containing life durations does not exist"
        self.pool_of_life_durations = ages_at_death
        self.common_parameters['life_expectancy'] = mean(ages_at_death)

        # read the country/region specific age_pyramids
        sheet_path = path.join('country_data', 'age_pyramids', 'formated_data.xlsx')
        if path.isfile(sheet_path):
            sheet = read_sheet(sheet_path)
            self.process_age_pyramid(sheet)
        else:
            print "Spreadsheet containing age pyramids does not exist"

        # read the contact calibration data by country
        sheet_path = path.join('country_data', 'age_preference', 'sd_agepref_work.xlsx')
        if path.isfile(sheet_path):
            sheet = read_sheet(sheet_path)
            self.sd_agepref_work = sheet_to_dict(sheet)
        else:
            print "Spreadsheet sd_agepref_work.xlsx does not exist"

        # total contact rate by age category and by contact_type using Prem data
        # also full matrix for all locations
        if self.console['country'] != "None":
            total_contact_rate_prem = {}
            for contact_type in ['school', 'work', 'other_locations']:
                matrix = contact_calibration.read_matrix(contact_type, self.console['country'])
                # adjust workplace contacts as not everyone is working
                if contact_type == 'work':
                    matrix /= self.common_parameters['perc_active']/100.
                total_contact_rate_prem[contact_type] = matrix.sum(axis=1)
                self.prem_contact_rate_functions[contact_type] = self.create_prem_contact_rate_functions(
                    total_contact_rate_prem[contact_type])
                self.contact_rates_matrices[contact_type] = matrix

        # load country-specific siler parameters
        if self.console['country'] != "None":
            iso3 = get_iso3(self.console['country'])
            sheet_path = path.join('country_data', 'ages_at_death', 'siler_params.xlsx')
            if path.isfile(sheet_path):
                sheet = read_sheet(sheet_path)
                for line in range(sheet.max_row + 1)[2:]:
                    if sheet.cell(row=line, column=1).value.encode("utf-8") == iso3:
                        for j in range(2, 7):
                            self.siler_params[sheet.cell(row=1, column=j).value.encode("utf-8")] =\
                                float(sheet.cell(row=line, column=j).value)
            else:
                print "Spreadsheet containing Siler parameters does not exist"

        # time-variant parameters:
        if self.console['country'] != "None":
            keys_of_sheet_to_read = ['bcg_2016', 'gtb_2015', 'gtb_2016', 'outcomes_2015']
            self.data_from_sheets = spreadsheet.read_input_data_xls(True, keys_of_sheet_to_read,
                                                                    self.console['country'], False)
            # manual updates
            self.data_from_sheets['outcomes']['c_new_tsr'].update({1950: 0.})
            self.data_from_sheets['gtb_2016']['c_cdr'].update({1950: 0.})
            if self.console['country'] == 'India':
                self.data_from_sheets['bcg'].update({1975: 0.})
                del self.data_from_sheets['outcomes']['c_new_tsr'][1994]
                del self.data_from_sheets['outcomes']['c_new_tsr'][1995]
                del self.data_from_sheets['outcomes']['c_new_tsr'][1996]

                self.data_from_sheets['outcomes']['c_new_tsr'].update({1980: 0.})
            elif self.console['country'] == 'Indonesia':
                del self.data_from_sheets['bcg'][1980]
                self.data_from_sheets['bcg'].update({1950: 0.})
                del self.data_from_sheets['gtb_2016']['c_cdr'][1950]
                self.data_from_sheets['gtb_2016']['c_cdr'][1980] = 0.
            elif self.console['country'] == 'China':
                self.data_from_sheets['bcg'][1970] = 0.
                del self.data_from_sheets['gtb_2016']['c_cdr'][1950]
                self.data_from_sheets['gtb_2016']['c_cdr'][1980] = 0.
            elif self.console['country'] == 'Philippines':
                del self.data_from_sheets['outcomes']['c_new_tsr'][1994]
                del self.data_from_sheets['outcomes']['c_new_tsr'][1996]
                self.data_from_sheets['outcomes']['c_new_tsr'][1970] = 0.
                self.data_from_sheets['outcomes']['c_new_tsr'][1990] = 30.
                self.data_from_sheets['bcg'][1950] = 0.
                del self.data_from_sheets['bcg'][1980]
                del self.data_from_sheets['bcg'][1981]
            elif self.console['country'] == 'Pakistan':
                self.data_from_sheets['bcg'][1970] = 0.
                del self.data_from_sheets['outcomes']['c_new_tsr'][1998]
                self.data_from_sheets['gtb_2016']['c_cdr'][1995] = 0.

    def calculate_scale_up_functions(self):
        if self.console['country'] == "None":  # no scale-up, just constant parameter values
            def bcg_coverage_func(time):
                return self.common_parameters['vaccine_coverage'] / 100.
            self.scale_up_functions['bcg_coverage_prop'] = bcg_coverage_func

            def treatment_success_func(time):
                return self.common_parameters['perc_treatment_success'] / 100.
            self.scale_up_functions['treatment_success_prop'] = treatment_success_func

            def cdr_func(time):
                return self.common_parameters['perc_cdr_smearpos'] / 100.
            self.scale_up_functions['cdr_prop'] = cdr_func
        else:
            datasets = {'bcg_coverage_prop': copy.deepcopy(self.data_from_sheets['bcg']), 'treatment_success_prop':
                copy.deepcopy(self.data_from_sheets['outcomes']['c_new_tsr']), 'cdr_prop': copy.deepcopy(self.data_from_sheets['gtb_2016']['c_cdr'])}

            # smear-positive proportion separately
            sp_number = self.data_from_sheets['outcomes']['new_sp_coh']
            snep_number = self.data_from_sheets['outcomes']['new_snep_coh']
            dataset_for_prop = {}
            for year in snep_number.keys():
                if year in sp_number.keys() and (snep_number[year] + sp_number[year]) > 0.:
                    dataset_for_prop[year] = 100.* sp_number[year] / (sp_number[year] + snep_number[year])

            datasets.update({'sp_prop': dataset_for_prop})

            for key, dataset in datasets.iteritems():
                # ignore some points for fitting
                if self.country == 'India' and key == 'cdr_prop':
                    del dataset[2000]
                    del dataset[2001]
                    # del dataset[2002]
                if self.country == 'Philippines' and key == 'cdr_prop':
                    del dataset[2000]

                x_vals = dataset.keys()
                y_vals = [val/100. for val in dataset.values()]  # perc to prop

                if key == 'treatment_success_prop':
                    bound_high = 0.95
                else:
                    bound_high = 1.
                fn = scale_up_function(x_vals, y_vals, 5, .1, bound_low=0., bound_up=bound_high)
                self.scale_up_functions[key] = deepcopy(fn)

    def read_country_parameters(self, sheet):
        col_index = None
        for col in range(sheet.max_column + 1)[2:]:
            country_name = sheet.cell(row=1, column=col).value
            if country_name == self.console['country']:
                col_index = col
                break
        if col_index is None:
            print "WARNING: country " + self.console['country'] + " was not found in the country_parameters.xlsx spreadsheet."
            return

        for row in range(sheet.max_row + 1)[2:]:
            param_name = sheet.cell(row=row, column=1).value
            param_value = sheet.cell(row=row, column=col_index).value
            if param_value is not None:
                if param_name in self.console.keys():
                    self.console[param_name] = param_value
                elif param_name in self.common_parameters.keys():
                    self.common_parameters[param_name] = sheet.cell(row=row, column=col_index).value
                else:
                    print "WARNING: parameter " + param_name + " was not found in console.xlsx or common_parameters.xlsx"
                    print "Its country-specifuc value for " + self.console['country'] + " was ignored."

    def process_age_pyramid(self, sheet):
        """
        Process the data found in the age_pyramid file
        """

        if self.console['country'] == "None":
            region = "World"
        else:
            region = self.console['country']
        for line in range(sheet.max_row + 1)[2:]:
            if sheet.cell(row=line, column=1).value == region:
                for col_index in range(4, 21):
                    cat_name = sheet.cell(row=1, column=col_index).value.encode("utf-8")
                    value = float(sheet.cell(row=line, column=col_index).value)
                    self.age_pyramid[cat_name] = value

        # normalise the vector so it sums to 1.0. It now contains proportions
        s = sum(self.age_pyramid.values())
        for key, value in self.age_pyramid.iteritems():
            self.age_pyramid[key] /= s

    def create_contact_rate_function(self):
        """
        create a function that generates contact rates according to age
        """
        def f(age):
            for i, bounds in enumerate(self.contact_rates['bounds']):
                if age < bounds[1]:
                    rate = self.contact_rates['contact_rate'][i]
                    break
            return rate

        self.contact_rate_function = f

    def create_prem_contact_rate_functions(self, rates):
        def f(age):
            if age >= 75.:
                prem_indice = 15
            else:
                prem_indice = int(floor(age/5.))
            return rates[prem_indice]
        return f

    def process_checkpoints(self):
        """
        the chekpoints are entered as a character string such as 't0/t1/t2.../tn' in the spreadsheets
        they need to be stored as the tuple [t0, t1, t2..., tn]
        """
        if self.console['years_checkpoints'] != 'None':
            processed_checkpoints = self.console['years_checkpoints'].split('/')
            processed_checkpoints = [int(ch) for ch in processed_checkpoints]
            self.console['years_checkpoints'] = processed_checkpoints
        else:
            self.console['years_checkpoints'] = []

    def workout_timeseries(self):
        """
        According to the outputs required from the console spreadsheet, define the list of timeseries to be recorded
        """
        self.console['timeseries_to_record'] = []
        for param in self.console.keys():
            if 'plot_ts_' in param:
                if self.console[param]:
                    self.console['timeseries_to_record'].append(param[8:])

        # add prop_mdr_prevalence
        if 'tb_prevalence_mdr' in  self.console['timeseries_to_record'] and 'tb_prevalence' in self.console['timeseries_to_record']:
            self.console['timeseries_to_record'].append('prop_mdr_prevalence')

    def workout_universal_methods(self):
        """
        According to which outputs are requested (plot, tables), determine whether universal methods are needed. A
        universal method is a method that requires to loop through each individual at each time-step.
        """
        self.console['run_universal_methods'] = False
        if self.console['plot_ts_mean_age'] or self.console['plot_ts_prop_under_5']:
            self.console['run_universal_methods'] = True
            print "Universal methods are requested. The running time will be increased"

    def workout_birth_rates(self):
        """
        Build a function that returns the average number of births that should be triggered at each time-step in order
        to match the age-pyramid at the end of the demographic burn-in.
        """
        def f(t):
            """
            :param t: time to end of demo-burning + tb_burning in days. i.e. age of the new-born individuals when the age-pyramid will be
            recorded
            :return: the average number of births per time-step
            """
            t /= 365.25 # t is now in years
            interval_width = 5.
            if t > 80.:
                interval_width = 20.  # going from 75yo to 100yo
            age_cat = get_agecategory(t, 'pyramid')
            nb_births = self.common_parameters['population'] * self.age_pyramid[age_cat] /\
                                  self.siler_survival_proba(t)
            # linear adjustment to smooth the 80+ category and prevent from having ades uniformly distributed within this category
            if t > 80.:
                nb_births *= 1.8 - (t-80.)*1.6/20.  # so that nb(100) = 0.2*nb_births  and nb(80) = 1.8*nb_birth

            nb_births *= self.console['time_step'] / (interval_width * 365.25)
            return nb_births
        self.birth_numbers_function = f

    def siler_survival_proba(self, x):
        """
        :param x: age in years!
        :return: survival probability
        """
        alpha1 = self.siler_params['alpha1']
        beta1 = self.siler_params['beta1']
        alpha2 = self.siler_params['alpha2']
        beta2 = self.siler_params['beta2']
        alpha3 = self.siler_params['alpha3']
        y = exp(
            ((exp(alpha1)) / beta1) * (exp(-beta1 * x) - 1) - ((exp(alpha2)) / beta2) * (exp(beta2 * x) - 1) - x * exp(
                alpha3))
        return y

    def draw_lhs_parameters(self):
        n_params = len(self.uncertainty_params)
        lhs_cube = lhs_sampler(n_params=n_params, n_samples=self.console['n_lhs_paramsets'])
        self.sampled_params = {}
        for i, param in enumerate(self.uncertainty_params.keys()):
            distrib = self.uncertainty_params[param]
            if distrib['distri'] == 'uniform':
                self.sampled_params[param] = uniform(distrib['pars'][0], distrib['pars'][1] - distrib['pars'][0]).ppf(lhs_cube[:, i])
            elif distrib['distri'] == 'triangular':
                self.sampled_params[param] = triang(loc=distrib['pars'][0], scale=distrib['pars'][1]-distrib['pars'][0],
                                                    c=0.5).ppf(lhs_cube[:, i])
            elif distrib['distri'] == 'beta':
                self.sampled_params[param] = beta(distrib['pars'][0], distrib['pars'][1]).ppf(lhs_cube[:, i])
            else:
                print distrib['distri'] + "distribution not supported."

    def write_lhs_parameters(self):
        n_params = len(self.uncertainty_params)
        a = zeros((self.console['n_lhs_paramsets'], n_params + 1))
        header = ''
        for j, param in enumerate(self.uncertainty_params.keys()):
            a[:, j] = self.sampled_params[param]
            if j > 0:
                header += ','
            header += param
        a[:, j+1] = range(self.console['n_lhs_paramsets'])

        header += ',sample_id'

        file_path = path.join('outputs', self.console['project_name'], 'lhs_values.csv')

        savetxt(file_path, a, delimiter=",", header=header)
